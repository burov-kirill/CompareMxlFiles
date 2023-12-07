import difflib
import filecmp
import glob
import hashlib
import os
import queue
import re
import sys
import threading
import time
import pandas as pd
from threading import Thread
import multiprocessing
import concurrent.futures as pool
import PySimpleGUI as sg
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from create_frame import create_dataframe_from_file
from interface import start, end, error
from logs import log
from queue import Empty
class ThreadWithReturnValue(Thread):

    def __init__(self, group=None, target=None, name=None,
                 args=(), kwargs={}, Verbose=None):
        Thread.__init__(self, group, target, name, args, kwargs)
        self._return = None
        self._stop_event = threading.Event()


    def run(self):
        if self._target is not None:
            self._return = self._target(*self._args,
                                        **self._kwargs)

    def stop(self):
        self._stop_event.set()

    def stopped(self):
        return self._stop_event.is_set()

    def join(self, *args):
        Thread.join(self, *args)
        return self._return

compare_dict = {True: 'Файлы идентичны', False: 'Файлы отличаются'}
FORMAT = '*.mxl'
RESULT_FILE_NAME = 'Результат сверки.xlsx'
# FULL_RESULT_FILE_NAME = str(pathlib.Path(__file__).parent.resolve()) + '\\' + RESULT_FILE_NAME
def get_files(path):
    file_list = glob.glob(os.path.join(path, "*.mxl"))
    folder = path[path.rfind('/')+1:]
    return file_list, folder

def create_files_dict(file_list1, direct_name1, file_list2, direct_name2):
    # Связать файлы
    files_dict = dict()
    for file in file_list1:
        short_name = get_short_name(file)
        match_file = [other_file for other_file in file_list2 if get_short_name(other_file) == short_name]
        if len(match_file) == 1:
            files_dict[file] = match_file[0]
    unmatched_files1 = [file for file in file_list1 if file not in files_dict.keys()]
    unmatched_files2 = [file for file in file_list2 if file not in files_dict.values()]
    if len(unmatched_files2) != 0 or len(unmatched_files1) != 0:
        str_list1 = "\n".join(unmatched_files1)
        str_list2 = "\n".join(unmatched_files2)
        # Записать в лог файл!!!
        unmatched_note1 = f'Для следующих файлов из папки {direct_name1} не были найдены соответствия в папке {direct_name2}:\n{str_list1}'
        unmatched_note2 = f'Для следующих файлов из папки {direct_name2} не были найдены соответствия в папке {direct_name1}:\n{str_list2}'
        # write_results(unmatched_note1, unmatched_note2)
    return files_dict

def compare_files_with_filecmp(first_file, second_file):
    # Сверка файлов с помощью filecmp
    compare_result = filecmp.cmp(first_file, second_file)
    notes_list = prepare_result_notes(first_file, second_file, compare_result)
    write_results(*notes_list)

def compare_only_data_in_files(first_file, second_file, is_single_file=False):
    COMPARE_REPORT = 'Подробный отчет.txt'
    first_data, second_data = parse_data(first_file), parse_data(second_file)
    first_hash, second_hash = calculate_hash(first_data), calculate_hash(second_data)
    compare_result = first_hash == second_hash
    result = prepare_result_notes(first_file, second_file, compare_result)
    if is_single_file:
        parse_df_f = create_dataframe_from_file(first_file)
        parse_df_s = create_dataframe_from_file(second_file)
        string_f = '\n'.join([' '.join([element for element in lst if element != ""]) for lst in parse_df_f.values])
        string_s = '\n'.join([' '.join([element for element in lst if element != ""]) for lst in parse_df_s.values])
        difference = difflib.Differ()
        with open(COMPARE_REPORT, 'w', encoding='utf-8') as wrt_file:
            for line in difference.compare(string_f.splitlines(keepends=True), string_s.splitlines(keepends=True)):
                wrt_file.write(line)
    return result

def calculate_hash(string_list):
    hash_value = hashlib.md5(''.join(string_list).encode()).hexdigest()
    return hash_value


def parse_data(file):
    pattern = '\{\"#\",(.*?)\}'
    with open(file, 'r', encoding='utf-8') as fl:
        raw_data = fl.read()
        data_list = re.findall(pattern, raw_data)
        return data_list


def get_compare_ratio(first_file, second_file, ratio_dict):
    with open(first_file, 'r', encoding='utf-8') as f1, open(second_file, 'r', encoding='utf-8') as f2:
        seq_mat = difflib.SequenceMatcher()
        file1 = f1.read()
        file2 = f2.read()
        seq_mat.set_seqs(file1, file2)
        ratio_dict['ratio_coeff'] = seq_mat.real_quick_ratio()

def prepare_result_notes(first_file, second_file, compare_result):
    result_dict = dict()
    ratio_dict = multiprocessing.Manager().dict()
    keys = ['БИТ.Финанс файл', 'БИТ.Строительство файл', 'Коэффициент подобия', 'Результат сверки']
    first_shrt_name, second_shrt_name = get_short_name(first_file), get_short_name(second_file)
    # ratio_note = 0.5
    ratio_task = multiprocessing.Process(target=get_compare_ratio, args=[first_file, second_file, ratio_dict])
    ratio_task.start()
    ratio_task.join(10)
    if ratio_task.is_alive():
        ratio_task.terminate()
    if ratio_dict.values() == []:
        ratio_note = 0.0
    else:
        ratio_note = round(ratio_dict['ratio_coeff'],2)
    for key, value in zip(keys, [first_shrt_name, second_shrt_name, ratio_note, compare_dict[compare_result]]):
        result_dict[key] = value
    return result_dict
def write_results(res_frame, save_path):
    # Записать результаты
    res_frame.to_excel(save_path, index=False)
    decorate_file(res_frame, save_path)

def decorate_file(res_frame, save_path):
    # Оформление файла
    INDENT = 5
    redFill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    greenFill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    COEFF_RANGE, RESULT_RANGE = f'C2:C{1 + len(res_frame)}', f'D2:D{1 + len(res_frame)}'

    red = Font(color='9C0006')
    red_dxf = DifferentialStyle(font=red, fill=redFill)
    red_rule = Rule(type="expression", formula=['NOT(ISERROR(SEARCH("Файлы отличаются",D2)))'], text="Файлы отличаются", dxf=red_dxf, stopIfTrue=True)

    green = Font(color='006100')
    green_dxf = DifferentialStyle(font=green, fill=greenFill)
    green_rule = Rule(type='expression',formula=['NOT(ISERROR(SEARCH("Файлы идентичны",D2)))'],  text="Файлы идентичны", dxf=green_dxf, stopIfTrue=True)
    workbook = load_workbook(filename=save_path)
    ws = workbook.active
    ws.title = 'Сверка файлов'
    table = Table(displayName="Table1", ref=f"A1:D{1 + len(res_frame)}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.conditional_formatting.add(COEFF_RANGE, ColorScaleRule(start_type='min', start_color='F8696B',
                                                            end_type = 'max', end_color = '63BE7B'))

    # ws.conditional_formatting.add(RESULT_RANGE, FormulaRule(formula=['NOT(ISERROR(SEARCH("Файлы идентичны",D2)))'], stopIfTrue=True,
    #                                                       fill=greenFill))
    # ws.conditional_formatting.add(RESULT_RANGE, FormulaRule(formula=['NOT(ISERROR(SEARCH("Файлы отличаются",D2)))'], stopIfTrue=True,
    #                                                       fill=redFill))

    ws.conditional_formatting.add(RESULT_RANGE, red_rule)
    ws.conditional_formatting.add(RESULT_RANGE, green_rule)
    for i, column in enumerate(res_frame.columns, 1):
        data_list = [len(str(value)) for value in res_frame[column]]
        data_list.append(len(column))
        max_width = max(data_list) + INDENT
        ws.column_dimensions[get_column_letter(i)].width = max_width

    for i in range(1, len(res_frame)+2):
        ws.cell(row=i, column=3).number_format = '0.00%'
    workbook.save(save_path)


def get_short_name(abs_path):
    return os.path.basename(abs_path)

def main_task(files_dict, is_single):
    global q
    result_list = []
    counter = 0
    for k, v in files_dict.items():
        counter+=1
        short_file_name = get_short_name(k)
        q.put(short_file_name)
        try:
            res = compare_only_data_in_files(k, v, is_single)
        except Exception as exp:
            # Запись лога исключения
            log.info(f'При обработке следующих файлов {k}, {v} возникло исключение:\n')
            log.exception(exp)
        else:
            result_list.append(res)
            # Запись лога успешной обработки
            log.info(f'Обработка следующих файлов {k}, {v} успешно завершена\n')
        finally:
            q.put(counter)
    return result_list

if __name__ == '__main__':
    multiprocessing.freeze_support()
    values = start()
    fin_folder, build_folder, save_folder = values['fin_ipt'], values['build_ipt'], values['save_folder']
    is_single_file = values['is_single_file']
    save_path = save_folder + '\\' + RESULT_FILE_NAME
    if not is_single_file:
        fin_files, fin_folder_name = get_files(fin_folder)
        build_files, build_folder_name = get_files(build_folder)
        files_dict = create_files_dict(fin_files, fin_folder_name,  build_files, build_folder_name)
    else:
        files_dict = {fin_folder:build_folder}

    # Инициализация прогрессбара
    progressbar = [[sg.ProgressBar(len(files_dict), size=(100, 20),  orientation='h', key='pg_bar')]]
    outputwin = [[sg.Output(key='out', size=(100, 20))]]
    layout = [
        [sg.Frame('Прогресс', layout=progressbar, background_color='#007bfb', size=(400, 50), key='prg_frame')],
        [sg.Frame('Файл', layout=outputwin,  background_color='#007bfb', size=(400, 50))]
    ]
    window = sg.Window('Работа', layout=layout, finalize=True, element_justification='center', background_color='#007bfb')
    pg_bar = window['pg_bar']
    out = window['out']
    q = queue.Queue()
    worker_task = ThreadWithReturnValue(target=main_task, args=[files_dict, is_single_file])
    worker_task.setDaemon(True)
    worker_task.start()
    while True:
        event, values = window.read(timeout=100)
        if event == 'Cancel' or event is None:
            sys.exit()
        try:
            value = q.get_nowait()
        except Empty:
            continue
        else:
            if isinstance(value, int):
                pg_bar.UpdateBar(value)
                window.Element('prg_frame').Update(f"{value} из {len(files_dict)}")
                if value == len(files_dict):  #
                    break
            else:
                window.Element('out').Update(value)
    window.close()

    compare_result = worker_task.join()
    try:
        result_frame = pd.DataFrame(compare_result)
        write_results(result_frame, save_path)
    except Exception as exp:
        log.info('!Возникла непредвиденная ошибка')
        log.exception(exp)
        error()
    else:
        end(save_path)



